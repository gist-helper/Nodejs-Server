import openpyxl
import json
from datetime import datetime

from meal import MealWrapper, kind_of_meals, kind_of_meals_eng, slot_endpoints, slot_filenames_postfix, slot_filenames_postfix_eng

from util import ComplexEncoder, sanitize_menu

file_path = "./2학생회관.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.worksheets[0]

for column_day in range(4,11):
    for row_slot in range(3):
        day = sheet.cell(row=2, column=column_day).value

        menus = ""
        for row_menus in range(slot_endpoints[row_slot][0], slot_endpoints[row_slot][1]):
            menus += sanitize_menu((sheet.cell(row=row_menus, column=column_day).value)) + '\n'
            if row_menus == 20:
                menus += "\n\\코너\\\n"

        meal_wrapper = MealWrapper()
        meal_wrapper.meal.title="제2학생회관1층"
        try :
            day.strftime('%Y-%m-%d')
            # mealString = mealDate.strftime('%m_%d')
        except AttributeError:
            day = datetime(2022, int(day[0:2]), int(day[4:6]), 00, 00, 00)
            # mealString = mealDate.strftime('%m_%d')
        meal_wrapper.meal.meal_date = day.strftime('%Y-%m-%d')
        meal_wrapper.meal.kind_of_meal = kind_of_meals[row_slot]
        meal_wrapper.meal.menu = menus.rstrip('\n')

        default_name = day.strftime('%m_%d') + slot_filenames_postfix[row_slot]
        jsonFile = open(f"./{default_name}", 'w+',encoding='utf-8')
        json.dump(meal_wrapper.__dict__,jsonFile,indent=4,ensure_ascii=False,cls=ComplexEncoder)

sheet = workbook.worksheets[1]

for column_day in range(4,11):
    for row_slot in range(3):
        day = sheet.cell(row=2, column=column_day).value

        menus = ""
        for row_menus in range(slot_endpoints[row_slot][0], slot_endpoints[row_slot][1]):
            menus += sanitize_menu((sheet.cell(row=row_menus, column=column_day).value)) + '\n'
            if row_menus == 20:
                menus += "\n\\Corner\\\n"

        meal_wrapper = MealWrapper()
        meal_wrapper.meal.title="Student Union Bldg.2 1st floor"
        try :
            day.strftime('%Y-%m-%d')
            # mealString = mealDate.strftime('%m_%d')
        except AttributeError:
            day = datetime(2022, int(day[0:2]), int(day[4:6]), 00, 00, 00)
            # mealString = mealDate.strftime('%m_%d')
        meal_wrapper.meal.meal_date = day.strftime('%Y-%m-%d')
        meal_wrapper.meal.kind_of_meal = kind_of_meals_eng[row_slot]
        meal_wrapper.meal.menu = menus.rstrip('\n')

        default_name = day.strftime('%m_%d') + slot_filenames_postfix_eng[row_slot]
        jsonFile = open(f"./{default_name}", 'w+',encoding='utf-8')
        json.dump(meal_wrapper.__dict__,jsonFile,indent=4,ensure_ascii=False,cls=ComplexEncoder)
