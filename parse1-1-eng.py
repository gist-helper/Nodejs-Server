import openpyxl
import codecs
import datetime


def replaceToJson(text):
    if text == None:
        return ""
    else:
        text = text.replace("\n", "\\n")
        text = text.replace('"', '\\"')
        return text


excel_path = "./1학생회관_E.xlsx"
wb = openpyxl.load_workbook(excel_path)
sh = wb.worksheets[1]

date_excel_path = "./2학생회관.xlsx"
date_wb = openpyxl.load_workbook(date_excel_path)
date_sh = date_wb.worksheets[0]

mealDate = date_sh.cell(row=2, column=4).value
print(mealDate)
print(type(mealDate))
mealDate += datetime.timedelta(days=-1)

for i in range(3, 13,2):
    mealDate += datetime.timedelta(days=1)
    for j in range(2):
        jsonFile = codecs.open(f'{10+(i-3)+j}.json', 'w+', 'utf-8')

        jsonFile.write('{\n\t"meal":{' + '\n')

        jsonFile.write('\t\t"title": "Student Union Bldg. 1 1st floor",\n')

        jsonFile.write('\t\t' + '"meal_date": ' + '"%s"' %
                       (mealDate.strftime('%Y-%m-%d')) + ",\n")

        jsonword = ""
        if j == 0:
            jsonFile.write('\t\t' + '"kind_of_meal": "Lunch",\n')

            jsonFile.write('\t\t' + '"menu": ')
            jsonword += "Original\\n"
            for k in range(6, 13):
                print("k=",k)
                print("i=",i)
                jsonword += (
                    replaceToJson(sh.cell(row=k, column=i).value) + " " +'\\n')
                if k == 10:
                    jsonword += "\\nCorner\\n"
        else:
            jsonFile.write('\t\t' + '"kind_of_meal": "Dinner",\n')

            jsonFile.write('\t\t' + '"menu": ')
            for k in range(13, 20):
                jsonword += (
                    replaceToJson(sh.cell(row=k, column=i).value) + " " +
                     '\\n')
        jsonFile.write('"%s"' % (jsonword))
        jsonFile.write('\n\t}\n')

        jsonFile.write('}')
        jsonFile.close()
