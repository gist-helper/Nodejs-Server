import openpyxl
import codecs
from datetime import datetime


def replaceToJson(text):
    if text == None:
        return ""
    else:
        text = text.replace("\n", "\\n")
        text = text.replace('"', '\\"')
        return text


excel_path = "./2학생회관.xlsx"

wb = openpyxl.load_workbook(excel_path)
sh = wb.worksheets[1]
sh1 = wb.worksheets[0]

for i in range(4, 11):
    for j in range(3):
        mealDate = sh.cell(row=2, column=i).value
        try :
            mealDate.strftime('%Y-%m-%d')
            mealString = mealDate.strftime('%m_%d')
        except AttributeError:
            mealDate = datetime(2022, int(mealDate[0:2]), int(mealDate[4:6]), 00, 00, 00)
            mealString = mealDate.strftime('%m_%d')
        if j == 0:
            jsonFile = codecs.open(f'{mealString}_b_eng.json', 'w+', 'utf-8')
        elif j == 1:
            jsonFile = codecs.open(f'{mealString}_l_eng.json', 'w+', 'utf-8')
        else:
            jsonFile = codecs.open(f'{mealString}_d_eng.json', 'w+', 'utf-8')

        #jsonFile = codecs.open(f'{30+3*(i-4)+j}.json', 'w+', 'utf-8')

        jsonFile.write('{\n\t"meal":{' + '\n')

        jsonFile.write('\t\t"title": "Student Union Bldg.2 1st floor",\n')

        jsonFile.write('\t\t' + '"meal_date": ' + '"%s"' %
                       (mealDate.strftime('%Y-%m-%d')) + ",\n")

        jsonword = ""
        if j == 0:
            jsonFile.write('\t\t' + '"kind_of_meal": "Breakfast",\n')

            jsonFile.write('\t\t' + '"menu": ')
            for k in range(3, 13):
                jsonword += (replaceToJson(sh.cell(row=k, column=i).value) +
                             '\\n')
        elif j == 1:
            jsonFile.write('\t\t' + '"kind_of_meal": "Lunch",\n')

            jsonFile.write('\t\t' + '"menu": ')
            jsonword += "Original\\n"
            for k in range(13, 23):
                jsonword += (replaceToJson(sh.cell(row=k, column=i).value) +
                             '\\n')
                if k == 20:
                    jsonword += "\\nCorner\\n"
        else:
            jsonFile.write('\t\t' + '"kind_of_meal": "Dinner",\n')

            jsonFile.write('\t\t' + '"menu": ')
            for k in range(23, 30):
                jsonword += (replaceToJson(sh.cell(row=k, column=i).value) +
                             '\\n')
        jsonFile.write('"%s"' % (jsonword))
        jsonFile.write('\n\t}\n')

        jsonFile.write('}')
        jsonFile.close()
