import openpyxl
import json

from meal import MealWrapper, kind_of_meals, slot_endpoints, slot_filenames_postfix

from util import ComplexEncoder, sanitize_menu

# 1학생회관 1층 : 11월3째주메뉴.xlsx
# 1학생회관 2층 : 주간식단표(2022년11월21일~2022년11월25일식단표).xlsx
# 2학생회관 1층 : 11월3주.xlsx
file_path = "./2학생회관.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.worksheets[0]

for column_day in range(4,11):
    for row_slot in range(3):
        day = sheet.cell(row=2, column=column_day).value

        menus = ""
        for row_menus in range(slot_endpoints[row_slot][0], slot_endpoints[row_slot][1]):
            menus += sanitize_menu((sheet.cell(row=row_menus, column=column_day).value)) + '\n'
        menus += sanitize_menu((sheet.cell(row=slot_endpoints[row_slot][1], column=column_day).value))

        meal_wrapper = MealWrapper()
        meal_wrapper.meal.title="제2학생회관1층"
        meal_wrapper.meal.meal_date = day.strftime('%Y-%m-%d')
        meal_wrapper.meal.kind_of_meal = kind_of_meals[row_slot]
        meal_wrapper.meal.menu = menus

        default_name = day.strftime('%m_$d') + slot_filenames_postfix[row_slot]
        jsonFile = open(f"./{default_name}.json", 'w+',encoding='utf-8')
        json.dump(meal_wrapper.__dict__,jsonFile,indent=4,ensure_ascii=False,cls=ComplexEncoder)