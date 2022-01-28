import openpyxl
import codecs
import datetime


def replaceToJson(text):
    if text == None:
        return ""
    else:
        text = text.replace("\n", "\\n")
        text = text.replace('"', '\\"')
        return text


excel_path = "./1학생회관.xlsx"

wb = openpyxl.load_workbook(excel_path)
sh = wb.worksheets[0]

mealDate = sh.cell(row=5, column=3).value
print(mealDate)
print(type(mealDate))
mealDate += datetime.timedelta(days=-1)

for i in range(3, 17, 2):
    mealDate += datetime.timedelta(days=1)
    for j in range(3):
        jsonFile = codecs.open(f'{3*((i-3)//2)+j}.json', 'w+', 'utf-8')

        jsonFile.write('{\n\t"meal":{' + '\n')

        jsonFile.write('\t\t"title": "제1학생회관1층 Student Union Bldg. 1 1st floor",\n')

        jsonFile.write('\t\t' + '"meal_date": ' + '"%s"' %
                       (mealDate.strftime('%Y-%m-%d')) + ",\n")

        jsonword = ""
        if j == 0:
            jsonFile.write('\t\t' + '"kind_of_meal": "조식 breakfast",\n')

            jsonFile.write('\t\t' + '"menu": ')
            for k in range(7, 17):
                jsonword += (
                    replaceToJson(sh.cell(row=k, column=i).value) + " " +
                    replaceToJson(sh.cell(row=k, column=i + 1).value) + '\\n')
        elif j == 1:
            jsonFile.write('\t\t' + '"kind_of_meal": "중식 lunch",\n')

            jsonFile.write('\t\t' + '"menu": ')
            jsonword += "정식 original\\n"
            for k in range(17, 26):
                jsonword += (
                    replaceToJson(sh.cell(row=k, column=i).value) + " " +
                    replaceToJson(sh.cell(row=k, column=i + 1).value) + '\\n')
                if k == 23:
                    jsonword += "\\n코너 corner\\n"
        else:
            jsonFile.write('\t\t' + '"kind_of_meal": "석식 dinner",\n')

            jsonFile.write('\t\t' + '"menu": ')
            for k in range(26, 33):
                jsonword += (
                    replaceToJson(sh.cell(row=k, column=i).value) + " " +
                    replaceToJson(sh.cell(row=k, column=i + 1).value) + '\\n')
        jsonFile.write('"%s"' % (jsonword))
        jsonFile.write('\n\t}\n')

        jsonFile.write('}')
        jsonFile.close()
