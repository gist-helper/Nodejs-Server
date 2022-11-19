import openpyxl
import codecs
import re

def replaceToJson(text):
    if text == None:
        return ""
    else:
        text = text.replace("\n", "\\n")
        text = text.replace('"', '\\"')
        return text


excel_path = "./1학생회관_2.xlsx"

wb = openpyxl.load_workbook(excel_path)
sh = wb.worksheets[0]

for i in range(2, 7):
    jsonFile = codecs.open(f'{20+(i-2)}.json', 'w+', 'utf-8')
    jsonFile.write('{\n\t"meal":{' + '\n')
    jsonFile.write('\t\t"title": "제1학생회관2층",\n')
    mealDate = sh.cell(row=2, column=i).value
    jsonFile.write('\t\t' + '"meal_date": ' + '"%s"' %
                   (mealDate.strftime('%Y-%m-%d')) + ",\n")
    jsonword = ""
    jsonFile.write('\t\t' + '"kind_of_meal": "중식",\n')
    jsonFile.write('\t\t' + '"menu": ')
    for k in range(3, 13):
        if k == 5 or k == 7:
            continue
        else:
            jsonword += (re.sub('[^가-힣]','',replaceToJson(sh.cell(row=k, column=i).value)) +
                     '\\n')
    jsonFile.write('"%s"' % (jsonword))
    jsonFile.write('\n\t}\n')
    jsonFile.write('}')
    jsonFile.close()